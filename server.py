import csv
import json
import mimetypes
import os
import re
import subprocess
import sys
import threading
import io
from datetime import datetime, timedelta
from http.cookies import SimpleCookie
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import unquote, urlparse, parse_qs

# openpyxl for Excel reporting
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Imported email alert engine connection
import email_alerts
import database

BASE_DIR = Path(__file__).resolve().parent
WEB_DIR = BASE_DIR / "web"

HOST = "0.0.0.0"  # Adjusted to allow connections from other devices seamlessly
PORT = int(os.environ.get("PORT", 8080))  # Render (and most hosts) assign this at runtime

PING_TIMEOUT_SECONDS = 1
# How many days of second-by-second raw ping history to keep. Older days get
# permanently summarized (see database.rollup_and_prune_old_ping_logs) into
# the daily_summary table and the raw rows are deleted — this keeps the
# database small without losing week/month/custom report accuracy.
PING_LOG_RETENTION_DAYS = int(os.environ.get("PING_LOG_RETENTION_DAYS", 7))
# How long Minor Interruption rows are kept. These now back the "Minor"
# tab of the Incident Log (browsable and manually reclassifiable), not just
# the short Network Fluctuation rolling window, so they're kept for a long
# time rather than pruned after a couple of days.
MINOR_INTERRUPTION_RETENTION_DAYS = int(os.environ.get("MINOR_INTERRUPTION_RETENTION_DAYS", 180))

# --- Professional outage classification / fluctuation detection tuning ---
# Normal monitoring cadence: ping every device on its own independent
# schedule, all the time, regardless of whether anyone has the dashboard
# open in a browser (matches the "Ping checks every 5 seconds" behavior).
NORMAL_PING_INTERVAL_SECONDS = int(os.environ.get("NORMAL_PING_INTERVAL_SECONDS", 5))
# After the first failed ping, how many 1-second retries are made before an
# outage is confirmed. If any retry succeeds, the whole event is treated as
# Temporary Packet Loss (never becomes an outage at all).
VERIFICATION_RETRY_COUNT = int(os.environ.get("VERIFICATION_RETRY_COUNT", 3))
VERIFICATION_RETRY_INTERVAL_SECONDS = int(os.environ.get("VERIFICATION_RETRY_INTERVAL_SECONDS", 1))
# Once an outage is confirmed, ping every second until recovery.
OUTAGE_PING_INTERVAL_SECONDS = int(os.environ.get("OUTAGE_PING_INTERVAL_SECONDS", 1))
# Outages shorter than this are Minor Interruptions (diagnostics only, no
# report, no email, no availability impact). 60s and longer is a Verified
# Outage (saved, reported, reduces availability, triggers an email).
MINOR_INTERRUPTION_THRESHOLD_SECONDS = int(os.environ.get("MINOR_INTERRUPTION_THRESHOLD_SECONDS", 60))

# Network Fluctuation detection: 10+ Minor Interruptions for the same device
# within a rolling window triggers one email, then a cooldown before another
# fluctuation email can be sent for that same device.
FLUCTUATION_WINDOW_MINUTES = int(os.environ.get("FLUCTUATION_WINDOW_MINUTES", 30))
FLUCTUATION_THRESHOLD_COUNT = int(os.environ.get("FLUCTUATION_THRESHOLD_COUNT", 10))
FLUCTUATION_COOLDOWN_MINUTES = int(os.environ.get("FLUCTUATION_COOLDOWN_MINUTES", 30))

# How often the supervisor reconciles the live device list (picks up newly
# added devices, stops monitoring deleted ones). This is not the ping
# interval — each device is pinged on its own 5s/1s schedule by its own
# background thread; this just controls how quickly the roster is refreshed.
DEVICE_ROSTER_REFRESH_SECONDS = int(os.environ.get("MONITOR_INTERVAL_SECONDS", 10))

STATE_LOCK = threading.Lock()
MONITOR_STATE = {}
# device_id -> {"thread": Thread, "stop_event": threading.Event}
DEVICE_MONITOR_THREADS = {}
THREAD_REGISTRY_LOCK = threading.Lock()


def now_iso():
    return database.now_iso()


def make_id(prefix):
    return database.make_id(prefix)


def hash_password(password):
    return database.hash_password(password)


def verify_password(password, stored_hash):
    return database.verify_password(password, stored_hash)


def get_user_by_email(email):
    return database.get_user_by_email(email)


def public_user(user):
    return database.public_user(user)


def get_session_user(handler):
    cookie_header = handler.headers.get("Cookie", "")
    cookie = SimpleCookie(cookie_header)
    token_cookie = cookie.get("uptime_tools_session")

    if token_cookie:
        return database.get_user_for_session_token(token_cookie.value)

    return None


def user_companies(user_id):
    return database.user_companies(user_id)


def find_company(user_id, company_id):
    return database.find_company(user_id, company_id)


def company_devices(company_id):
    return database.company_devices(company_id)


def company_reports(company_id):
    return database.company_reports(company_id)


def clean_device(data, company_id, user_id):
    location = str(data.get("location", "")).strip()
    name = str(data.get("name", "")).strip()
    ip = str(data.get("ip", "")).strip()
    muted = 1 if data.get("muted") in (True, 1, "1", "true", "True") else 0

    if not ip:
        raise ValueError("IP address is required")

    return {
        "id": data.get("id") or make_id("dev"),
        "company_id": company_id,
        "user_id": user_id,
        "location": location or "-",
        "name": name or "Unnamed Device",
        "ip": ip,
        "muted": muted,
        "created_at": data.get("created_at") or now_iso(),
    }


def build_ping_command(ip):
    if os.name == "nt":
        return ["ping", "-n", "1", "-w", str(PING_TIMEOUT_SECONDS * 1000), ip]
    return ["ping", "-c", "1", "-W", str(PING_TIMEOUT_SECONDS), ip]


def parse_ping_time(output):
    match = re.search(r"time[=<]\s*(\d+(?:\.\d+)?)\s*ms", output, re.IGNORECASE)
    if not match:
        return "--"
    if "time<" in match.group(0).lower():
        return "<1 ms"
    return f"{match.group(1)} ms"


def _icmp_ping_available():
    """Check once at startup whether the system `ping` binary works here.

    Render's containers (and many other PaaS hosts) either don't ship the
    `ping` binary or block raw ICMP sockets, so `subprocess` ping always
    fails there even though the network itself is fine. We detect that once
    and fall back to a TCP-connect reachability check instead.
    """
    try:
        result = subprocess.run(
            build_ping_command("127.0.0.1"),
            capture_output=True,
            text=True,
            timeout=PING_TIMEOUT_SECONDS + 1,
        )
        return result.returncode == 0
    except Exception:
        return False


ICMP_PING_AVAILABLE = _icmp_ping_available()
if not ICMP_PING_AVAILABLE:
    print(
        "[ping] System 'ping' is unavailable or blocked in this environment "
        "(expected on Render's free tier). Falling back to TCP-connect checks "
        "on common ports (80, 443, 22, 3389, 8080)."
    )

TCP_FALLBACK_PORTS = (80, 443, 22, 3389, 8080)


def _tcp_reachable(ip):
    """Best-effort reachability check when ICMP isn't available.

    Tries a few common ports; if any of them accept a TCP connection (even a
    refused connection means something answered) we treat the host as
    ONLINE. This is a reasonable approximation of "is this host up", not a
    true ping, and won't report latency.
    """
    import errno
    import socket
    import time as _time

    # errno codes that mean "something on the network actively answered us"
    # even though the port itself refused the connection -- this still
    # proves the host is up, so it should count as ONLINE too.
    REFUSED_BUT_ALIVE = {errno.ECONNREFUSED}

    for port in TCP_FALLBACK_PORTS:
        start = _time.monotonic()
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(PING_TIMEOUT_SECONDS)
        try:
            code = sock.connect_ex((ip, port))
            elapsed_ms = (_time.monotonic() - start) * 1000
            # 0 = connected, and a fast "connection refused" also
            # proves the host is up and answering on the network.
            if code == 0 or code in REFUSED_BUT_ALIVE:
                return True, f"{elapsed_ms:.0f} ms"
        except Exception:
            pass
        finally:
            sock.close()
    return False, "--"


def _self_internet_reachable():
    """Quick check that OUR OWN network connection is up, by trying to reach
    a couple of independent, always-on public DNS resolvers. Used so a device
    outage that happened purely because the monitoring server's own internet
    connection dropped doesn't get blamed on the device -- it's classified
    Minor (cause='self_internet') regardless of how long it lasted."""
    import socket

    for ip in ("1.1.1.1", "8.8.8.8", "9.9.9.9"):
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(2)
        try:
            if sock.connect_ex((ip, 53)) == 0:
                return True
        except Exception:
            pass
        finally:
            sock.close()
    return False


def ping_device(device):
    if not ICMP_PING_AVAILABLE:
        online, ping = _tcp_reachable(device["ip"])
        return {**device, "raw_status": "ONLINE" if online else "OFFLINE", "ping": ping}

    try:
        result = subprocess.run(
            build_ping_command(device["ip"]),
            capture_output=True,
            text=True,
            timeout=PING_TIMEOUT_SECONDS + 1,
        )
    except Exception:
        return {**device, "raw_status": "OFFLINE", "ping": "--"}

    raw_status = "ONLINE" if result.returncode == 0 else "OFFLINE"
    ping = parse_ping_time(result.stdout) if raw_status == "ONLINE" else "--"
    return {**device, "raw_status": raw_status, "ping": ping}


def format_duration(start_iso, end_dt):
    start = datetime.fromisoformat(start_iso)
    return str(end_dt - start).split(".")[0]


def _default_monitor_state():
    return {
        "status": "UNKNOWN",           # confirmed status: UNKNOWN / ONLINE / OFFLINE
        "display_status": "UNKNOWN",   # what the dashboard shows, incl. "CHECK n/N" while verifying a failure
        "offline_start": None,         # ISO timestamp of the first failed ping of the *current* confirmed outage
        "down_count": 0,               # Verified Outages seen this run (Minor Interruptions never count here)
        "ping": "--",
        "fluctuation_alert_sent_at": None,
        "long_outage_alert_sent": False,  # whether the immediate "still offline" email has fired for this outage
        "self_internet_down_at_start": False,  # was OUR internet already down when this outage began?
    }


def resolve_target_receiver(device):
    company_match = database.find_company_by_id(device["company_id"])
    if company_match:
        return (
            company_match.get("receivers")
            or company_match.get("email")
            or email_alerts.EMAIL_CONFIG.get("receiver_email", "")
        )
    return email_alerts.EMAIL_CONFIG.get("receiver_email", "")


def record_online_check(device, raw, checked_at):
    """Normal 5-second monitoring tick where the ping succeeded. Recovery
    from a confirmed outage is handled by record_recovery() instead, since
    that path needs to classify the event first."""
    device_id = device["id"]
    with STATE_LOCK:
        state = MONITOR_STATE.setdefault(device_id, _default_monitor_state())
        state["status"] = "ONLINE"
        state["display_status"] = "ONLINE"
        state["offline_start"] = None
        state["ping"] = raw["ping"]

    database.insert_ping_log(
        device_id, device["user_id"], checked_at.isoformat(timespec="seconds"),
        "ONLINE", raw["ping"], "-",
    )


def record_verifying(device, attempt, ping_value):
    """Live-display-only update while a failure is being verified (1s
    retries in progress). Never touches down_count, ping_logs, reports, or
    email -- verification failures aren't outages yet."""
    device_id = device["id"]
    with STATE_LOCK:
        state = MONITOR_STATE.setdefault(device_id, _default_monitor_state())
        state["display_status"] = f"CHECK {attempt}/{VERIFICATION_RETRY_COUNT}"
        state["ping"] = ping_value


def record_temporary_packet_loss(device):
    """A retry succeeded before all VERIFICATION_RETRY_COUNT attempts were
    exhausted: Temporary Packet Loss. No outage, no availability impact, no
    report, no email -- simply resume normal 5-second monitoring."""
    device_id = device["id"]
    with STATE_LOCK:
        state = MONITOR_STATE.setdefault(device_id, _default_monitor_state())
        state["status"] = "ONLINE"
        state["display_status"] = "ONLINE"
        state["offline_start"] = None


def record_confirmed_outage_start(device, first_failure_dt):
    """Every verification retry also failed: this is now a confirmed
    outage. Per spec, the outage start time is the timestamp of the very
    first failed ping (not the end of the retry window). Classification
    into Minor Interruption vs Verified Outage happens later, at recovery,
    once the full duration is known -- as does the self-internet check."""
    device_id = device["id"]
    offline_start_iso = first_failure_dt.isoformat(timespec="seconds")
    self_down_at_start = not _self_internet_reachable()
    with STATE_LOCK:
        state = MONITOR_STATE.setdefault(device_id, _default_monitor_state())
        state["status"] = "OFFLINE"
        state["display_status"] = "OFFLINE"
        state["offline_start"] = offline_start_iso
        state["self_internet_down_at_start"] = self_down_at_start
        state["long_outage_alert_sent"] = False  # fresh outage -- allow one alert to fire again
    return offline_start_iso


def record_outage_still_down(device):
    device_id = device["id"]
    with STATE_LOCK:
        state = MONITOR_STATE.setdefault(device_id, _default_monitor_state())
        state["display_status"] = "OFFLINE"


def record_long_outage_alert(device, offline_start_iso, checked_at):
    """Fires the moment a confirmed outage has been down for at least
    MINOR_INTERRUPTION_THRESHOLD_SECONDS (60s) -- while the device is
    STILL offline, instead of waiting for it to recover. Sent at most
    once per outage (guarded by the long_outage_alert_sent flag, which
    gets reset the next time a fresh outage starts)."""
    device_id = device["id"]
    with STATE_LOCK:
        state = MONITOR_STATE.setdefault(device_id, _default_monitor_state())
        if state.get("long_outage_alert_sent"):
            return  # already sent for this outage
        state["long_outage_alert_sent"] = True

    if device.get("muted") in (True, 1, "1"):
        return

    target_receiver = resolve_target_receiver(device)
    device_info = {
        "name": device["name"],
        "location": device["location"],
        "time": datetime.fromisoformat(offline_start_iso),
    }
    try:
        email_alerts.send_long_outage_alert(device_info, device["ip"], checked_at, target_receiver)
    except Exception as e:
        print(f"[Email] Failed to send long-outage (still offline) alert: {e}")


def check_fluctuation(device, checked_at):
    """Network Fluctuation detection: FLUCTUATION_THRESHOLD_COUNT or more
    Minor Interruptions for this device within FLUCTUATION_WINDOW_MINUTES
    sends one email, then enforces a cooldown before another fluctuation
    email can be sent for the same device."""
    device_id = device["id"]
    since_iso = (checked_at - timedelta(minutes=FLUCTUATION_WINDOW_MINUTES)).isoformat(timespec="seconds")
    count = database.count_recent_minor_interruptions(device_id, since_iso)
    if count < FLUCTUATION_THRESHOLD_COUNT:
        return

    with STATE_LOCK:
        state = MONITOR_STATE.setdefault(device_id, _default_monitor_state())
        last_sent = state.get("fluctuation_alert_sent_at")
        if last_sent:
            elapsed_minutes = (checked_at - datetime.fromisoformat(last_sent)).total_seconds() / 60
            if elapsed_minutes < FLUCTUATION_COOLDOWN_MINUTES:
                return  # still in cooldown for this device
        state["fluctuation_alert_sent_at"] = checked_at.isoformat(timespec="seconds")

    if device.get("muted") in (True, 1, "1"):
        return

    target_receiver = resolve_target_receiver(device)
    try:
        email_alerts.send_network_fluctuation_alert(device, count, FLUCTUATION_WINDOW_MINUTES, target_receiver)
    except Exception as e:
        print(f"[Email] Failed to send fluctuation alert: {e}")


def record_recovery(device, offline_start_iso, checked_at, raw):
    """The device just answered a ping again after a confirmed outage. This
    is where Minor Interruption vs Verified Outage is decided, since only
    now is the full outage duration known. An outage is classified Minor
    if it was short (< MINOR_INTERRUPTION_THRESHOLD_SECONDS) OR if our own
    internet connection was down at the start or end of it -- in that case
    we can't blame the device, no matter how long it lasted."""
    device_id = device["id"]
    duration_seconds = (checked_at - datetime.fromisoformat(offline_start_iso)).total_seconds()
    is_muted = device.get("muted") in (True, 1, "1")

    with STATE_LOCK:
        state = MONITOR_STATE.setdefault(device_id, _default_monitor_state())
        self_down_at_start = state.get("self_internet_down_at_start", False)

    self_down_now = not _self_internet_reachable()
    caused_by_self_internet = bool(self_down_at_start or self_down_now)

    if duration_seconds < MINOR_INTERRUPTION_THRESHOLD_SECONDS or caused_by_self_internet:
        # --- Minor Interruption: shown in the Minor incident log only,
        #     never affects reports/outage_history/availability/exports ---
        cause = "self_internet" if caused_by_self_internet else "short"
        database.insert_minor_interruption({
            "id": make_id("mint"),
            "device_id": device_id,
            "user_id": device["user_id"],
            "company_id": device["company_id"],
            "device_name": device["name"],
            "ip_address": device["ip"],
            "location": device["location"],
            "start_time": offline_start_iso,
            "end_time": checked_at.isoformat(timespec="seconds"),
            "duration_seconds": round(duration_seconds, 1),
            "cause": cause,
            "created_at": now_iso(),
        })
        check_fluctuation(device, checked_at)
    else:
        # --- Verified Outage: saved, reported, reduces availability, emailed ---
        downtime_display = format_duration(offline_start_iso, checked_at)
        report = {
            "id": make_id("rep"),
            "company_id": device["company_id"],
            "device_id": device_id,
            "user_id": device["user_id"],
            "date": datetime.fromisoformat(offline_start_iso).strftime("%d-%m-%Y"),
            "location": device["location"],
            "name": device["name"],
            "ip": device["ip"],
            "offline": datetime.fromisoformat(offline_start_iso).strftime("%H:%M:%S"),
            "online": checked_at.strftime("%H:%M:%S"),
            "downtime": downtime_display,
            "created_at": now_iso(),
            # Precise timestamps for the permanent outage_history table.
            "offline_start_iso": offline_start_iso,
            "online_iso": checked_at.isoformat(timespec="seconds"),
            "duration_seconds": duration_seconds,
        }
        database.insert_reports([report])

        # Bracket the outage in the raw ping log with exactly the two rows
        # that matter (the confirmed first failure, and the recovery). The
        # existing analytics/report code only needs a leading OFFLINE row
        # followed by an ONLINE row to compute outage periods correctly --
        # and, per spec, this is the only OFFLINE evidence a Verified
        # Outage leaves behind, so Minor Interruptions never pollute
        # availability, total outage time, or outage count.
        database.insert_ping_log(
            device_id, device["user_id"], offline_start_iso, "OFFLINE", "--", "-",
        )
        database.insert_ping_log(
            device_id, device["user_id"], checked_at.isoformat(timespec="seconds"),
            "ONLINE", raw["ping"], downtime_display,
        )

        with STATE_LOCK:
            state = MONITOR_STATE.setdefault(device_id, _default_monitor_state())
            state["down_count"] += 1

        if not is_muted:
            target_receiver = resolve_target_receiver(device)
            outage_data = {
                "name": device["name"],
                "ip": device["ip"],
                "location": device["location"],
                "offline_time": datetime.fromisoformat(offline_start_iso),
                "online_time": checked_at,
                "downtime": downtime_display,
            }
            try:
                email_alerts.send_short_outage_report(outage_data, target_receiver)
            except Exception as e:
                print(f"[Email] Failed to send recovery report: {e}")

    with STATE_LOCK:
        state = MONITOR_STATE.setdefault(device_id, _default_monitor_state())
        state["status"] = "ONLINE"
        state["display_status"] = "ONLINE"
        state["offline_start"] = None
        state["ping"] = raw["ping"]


def get_live_status(company_id):
    """Read-only snapshot for the dashboard/API. Pinging itself happens
    continuously in the background (see device_monitor_loop), on each
    device's own independent 5-second/1-second schedule -- completely
    decoupled from this endpoint. Calling this repeatedly (e.g. from browser
    polling) never triggers extra pings, double-counts failures, or races
    with the background monitor, since exactly one thread ever pings a
    given device."""
    devices = company_devices(company_id)
    if not devices:
        return []

    results = []
    now = datetime.now()
    with STATE_LOCK:
        for device in devices:
            state = MONITOR_STATE.get(device["id"]) or _default_monitor_state()
            down_for = "-"
            if state["status"] == "OFFLINE" and state.get("offline_start"):
                down_for = format_duration(state["offline_start"], now)
            results.append({
                "id": device["id"],
                "company_id": device["company_id"],
                "location": device["location"],
                "name": device["name"],
                "ip": device["ip"],
                "status": state["display_status"],
                "confirmedStatus": state["status"],
                "ping": state.get("ping", "--"),
                "downCount": state.get("down_count", 0),
                "downFor": down_for,
                "muted": device.get("muted", 0),
            })

    return sorted(results, key=lambda item: (item["status"] != "OFFLINE", item["location"], item["name"]))


def parse_csv_devices(csv_text, company_id, user_id):
    # Excel/Notepad on Windows saves CSVs with a UTF-8 BOM prefix, which
    # corrupts the first header's name (e.g. "\ufeffLocation") and made
    # every row silently fail to match — that was the real "import isn't
    # working" bug.
    if csv_text.startswith("\ufeff"):
        csv_text = csv_text[1:]

    rows = list(csv.DictReader(csv_text.splitlines()))
    devices = []
    skipped = 0
    for row in rows:
        # Normalize header names (case/space-insensitive) so "LOCATION",
        # "Location ", "location" etc. all match.
        normalized = {}
        for k, v in row.items():
            if k is None:
                continue
            key = k.strip().lower().lstrip("\ufeff")
            normalized[key] = (v or "").strip()

        if not any(normalized.values()):
            continue  # fully blank row (common trailing newline in exported CSVs)

        location = normalized.get("location")
        name = normalized.get("name") or normalized.get("device name") or normalized.get("resource name")
        ip = normalized.get("ip") or normalized.get("ip address") or normalized.get("target host")

        try:
            devices.append(clean_device({"location": location, "name": name, "ip": ip}, company_id, user_id))
        except ValueError:
            # Don't let one bad/missing-IP row abort the entire import —
            # skip it and keep going, then report the count back.
            skipped += 1
            continue

    return devices, skipped


def parse_timeframe(timeframe, start_str, end_str):
    now = datetime.now()
    if timeframe == "today":
        start_dt = datetime(now.year, now.month, now.day)
        end_dt = now
    elif timeframe == "week":
        start_dt = now - timedelta(days=7)
        end_dt = now
    elif timeframe == "month":
        start_dt = now - timedelta(days=30)
        end_dt = now
    elif timeframe == "6months":
        start_dt = now - timedelta(days=180)
        end_dt = now
    elif timeframe == "custom":
        try:
            start_dt = datetime.fromisoformat(start_str.split("T")[0])
            end_dt = datetime.fromisoformat(end_str.split("T")[0]).replace(hour=23, minute=59, second=59)
        except Exception:
            start_dt = now - timedelta(days=7)
            end_dt = now
    else:
        start_dt = now - timedelta(days=7)
        end_dt = now
    return start_dt, end_dt


def clean_sheet_title(name, location):
    title = f"{name} ({location})"
    title = re.sub(r"[\\/\?\*:\[\]]", "", title)
    if len(title) > 31:
        title = title[:28] + "..."
    return title


def aggregate_pings(rows, timeframe, start_dt, end_dt):
    """Aggregate ping rows into chart-friendly buckets.
    Returns labels, uptime %, overall average, and chart type hint."""

    if timeframe == "today":
        chart_type = "bar"
        buckets = ["12 AM - 6 AM", "6 AM - 12 PM", "12 PM - 6 PM", "6 PM - 12 AM"]
        bucket_data = {b: {"total": 0, "online": 0, "latencies": []} for b in buckets}
        for r in rows:
            try:
                ts = datetime.fromisoformat(r["timestamp"])
                h = ts.hour
                if 0 <= h < 6:
                    key = "12 AM - 6 AM"
                elif 6 <= h < 12:
                    key = "6 AM - 12 PM"
                elif 12 <= h < 18:
                    key = "12 PM - 6 PM"
                else:
                    key = "6 PM - 12 AM"
                
                bucket_data[key]["total"] += 1
                if r["status"] == "ONLINE":
                    bucket_data[key]["online"] += 1
                lat = r["latency"]
                if lat and lat != "--":
                    nums = re.findall(r"[\d\.]+", lat)
                    if nums:
                        bucket_data[key]["latencies"].append(float(nums[0]))
            except Exception:
                continue
                
        labels = buckets
        uptime_data = []
        latency_data = []
        for b in buckets:
            d = bucket_data[b]
            uptime_pct = round((d["online"] / d["total"]) * 100) if d["total"] > 0 else None
            avg_lat = round(sum(d["latencies"]) / len(d["latencies"]), 1) if d["latencies"] else None
            uptime_data.append(uptime_pct)
            latency_data.append(avg_lat)

    elif timeframe == "week":
        chart_type = "bar"
        group_format = "%Y-%m-%d"
        display_format = "%b %d"  # e.g., Jul 03
        buckets = []
        curr = datetime(start_dt.year, start_dt.month, start_dt.day)
        while curr.date() <= end_dt.date():
            buckets.append(curr.strftime(group_format))
            curr += timedelta(days=1)
            
        bucket_data = {b: {"total": 0, "online": 0, "latencies": []} for b in buckets}
        for r in rows:
            try:
                ts = datetime.fromisoformat(r["timestamp"])
                key = ts.strftime(group_format)
                if key in bucket_data:
                    bucket_data[key]["total"] += 1
                    if r["status"] == "ONLINE":
                        bucket_data[key]["online"] += 1
                    lat = r["latency"]
                    if lat and lat != "--":
                        nums = re.findall(r"[\d\.]+", lat)
                        if nums:
                            bucket_data[key]["latencies"].append(float(nums[0]))
            except Exception:
                continue
                
        labels = []
        uptime_data = []
        latency_data = []
        for b in buckets:
            try:
                dt = datetime.strptime(b, group_format)
                labels.append(dt.strftime(display_format))
            except Exception:
                labels.append(b)
            d = bucket_data[b]
            uptime_pct = round((d["online"] / d["total"]) * 100) if d["total"] > 0 else None
            avg_lat = round(sum(d["latencies"]) / len(d["latencies"]), 1) if d["latencies"] else None
            uptime_data.append(uptime_pct)
            latency_data.append(avg_lat)

    elif timeframe == "month":
        chart_type = "bar"
        buckets = ["Week 1", "Week 2", "Week 3", "Week 4", "Week 5"]
        bucket_data = {b: {"total": 0, "online": 0, "latencies": []} for b in buckets}
        
        start_date_only = start_dt.date()
        for r in rows:
            try:
                ts = datetime.fromisoformat(r["timestamp"])
                delta_days = (ts.date() - start_date_only).days
                if delta_days < 7:
                    key = "Week 1"
                elif delta_days < 14:
                    key = "Week 2"
                elif delta_days < 21:
                    key = "Week 3"
                elif delta_days < 28:
                    key = "Week 4"
                else:
                    key = "Week 5"
                    
                bucket_data[key]["total"] += 1
                if r["status"] == "ONLINE":
                    bucket_data[key]["online"] += 1
                lat = r["latency"]
                if lat and lat != "--":
                    nums = re.findall(r"[\d\.]+", lat)
                    if nums:
                        bucket_data[key]["latencies"].append(float(nums[0]))
            except Exception:
                continue
                
        labels = buckets
        uptime_data = []
        latency_data = []
        for b in buckets:
            d = bucket_data[b]
            uptime_pct = round((d["online"] / d["total"]) * 100) if d["total"] > 0 else None
            avg_lat = round(sum(d["latencies"]) / len(d["latencies"]), 1) if d["latencies"] else None
            uptime_data.append(uptime_pct)
            latency_data.append(avg_lat)

    else:
        chart_type = "bar"
        group_format = "%Y-%m-%d"
        display_format = "%b %d"
        buckets = []
        curr = datetime(start_dt.year, start_dt.month, start_dt.day)
        while curr.date() <= end_dt.date():
            buckets.append(curr.strftime(group_format))
            curr += timedelta(days=1)
            
        bucket_data = {b: {"total": 0, "online": 0, "latencies": []} for b in buckets}
        for r in rows:
            try:
                ts = datetime.fromisoformat(r["timestamp"])
                key = ts.strftime(group_format)
                if key in bucket_data:
                    bucket_data[key]["total"] += 1
                    if r["status"] == "ONLINE":
                        bucket_data[key]["online"] += 1
                    lat = r["latency"]
                    if lat and lat != "--":
                        nums = re.findall(r"[\d\.]+", lat)
                        if nums:
                            bucket_data[key]["latencies"].append(float(nums[0]))
            except Exception:
                continue
                
        labels = []
        uptime_data = []
        latency_data = []
        for b in buckets:
            try:
                dt = datetime.strptime(b, group_format)
                labels.append(dt.strftime(display_format))
            except Exception:
                labels.append(b)
            d = bucket_data[b]
            uptime_pct = round((d["online"] / d["total"]) * 100) if d["total"] > 0 else None
            avg_lat = round(sum(d["latencies"]) / len(d["latencies"]), 1) if d["latencies"] else None
            uptime_data.append(uptime_pct)
            latency_data.append(avg_lat)

    # --- Overall average across all pings for true running live average ---
    total_pings = len(rows)
    online_pings = sum(1 for r in rows if r["status"] == "ONLINE")
    overall_avg = round((online_pings / total_pings) * 100) if total_pings > 0 else 100

    return {
        "labels": labels,
        "uptime": uptime_data,
        "latency": latency_data,
        "chartType": chart_type,
        "overallAvg": overall_avg,
    }


class Handler(BaseHTTPRequestHandler):
    timeout = 30  # don't let a stalled/slow connection block a thread forever

    def end_headers(self):
        self.send_header("Access-Control-Allow-Origin", self.headers.get("Origin", "*"))
        self.send_header("Access-Control-Allow-Credentials", "true")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, PUT, DELETE, OPTIONS")
        super().end_headers()

    def do_OPTIONS(self):
        self.send_response(204)
        self.end_headers()

    def do_GET(self):
        try:
            self._do_GET_inner()
        except (BrokenPipeError, ConnectionResetError):
            pass
        except Exception as exc:
            print(f"[error] GET {self.path} -> {exc}", file=sys.stderr)
            try:
                self.send_json({"error": "Something went wrong on the server. Please try again."}, 500)
            except Exception:
                pass

    def _do_GET_inner(self):
        parsed = urlparse(self.path)
        path = parsed.path
        query = parse_qs(parsed.query)

        if path == "/api/me":
            user = get_session_user(self)
            if not user:
                self.send_json({"authenticated": False})
                return
            self.send_json({"authenticated": True, "user": public_user(user)})
            return

        if path == "/api/companies":
            user = self.require_user()
            if not user:
                return
            self.send_json({"companies": user_companies(user["id"])})
            return

        match = re.fullmatch(r"/api/companies/([^/]+)/devices", path)
        if match:
            user = self.require_user()
            if not user:
                return
            company_id = unquote(match.group(1))
            if not find_company(user["id"], company_id):
                self.send_json({"error": "Company not found"}, 404)
                return
            self.send_json({"devices": company_devices(company_id)})
            return

        match = re.fullmatch(r"/api/companies/([^/]+)/reports", path)
        if match:
            user = self.require_user()
            if not user:
                return
            company_id = unquote(match.group(1))
            if not find_company(user["id"], company_id):
                self.send_json({"error": "Company not found"}, 404)
                return
            date_str = query.get("date", [""])[0]
            start_str = query.get("start", [""])[0]
            end_str = query.get("end", [""])[0]
            reports = database.company_reports(
                company_id,
                date_iso=date_str or None,
                start_date=start_str or None,
                end_date=end_str or None,
            )
            self.send_json({"reports": reports})
            return

        # ── Incident Log: Major (Verified Outages) or Minor (short blips /
        # our-own-internet drops / manually demoted) incidents, same shape ──
        match = re.fullmatch(r"/api/companies/([^/]+)/incidents", path)
        if match:
            user = self.require_user()
            if not user:
                return
            company_id = unquote(match.group(1))
            if not find_company(user["id"], company_id):
                self.send_json({"error": "Company not found"}, 404)
                return
            incident_type = (query.get("type", ["major"])[0] or "major").lower()
            start_str = query.get("start", [""])[0]
            end_str = query.get("end", [""])[0]
            device_id_filter = query.get("device_id", [""])[0]

            if incident_type == "minor":
                incidents = database.list_minor_interruptions(
                    company_id,
                    device_id=device_id_filter or None,
                    start_date=start_str or None,
                    end_date=end_str or None,
                )
            else:
                incident_type = "major"
                incidents = database.company_reports(
                    company_id,
                    start_date=start_str or None,
                    end_date=end_str or None,
                )
                if device_id_filter:
                    incidents = [r for r in incidents if r.get("device_id") == device_id_filter]

            self.send_json({"incidents": incidents, "type": incident_type})
            return

        match = re.fullmatch(r"/api/companies/([^/]+)/status", path)
        if match:
            user = self.require_user()
            if not user:
                return
            company_id = unquote(match.group(1))
            if not find_company(user["id"], company_id):
                self.send_json({"error": "Company not found"}, 404)
                return
            self.send_json({"devices": get_live_status(company_id)})
            return

        # NEW HISTORICAL ANALYTICS ENDPOINT
        match = re.fullmatch(r"/api/devices/([^/]+)/analytics", path)
        if match:
            user = self.require_user()
            if not user:
                return
            device_id = unquote(match.group(1))
            device = database.get_device_for_user(device_id, user["id"])
            if not device:
                self.send_json({"error": "Device not found"}, 404)
                return
                
            timeframe = query.get("timeframe", ["week"])[0]
            start_str = query.get("start", [""])[0]
            end_str = query.get("end", [""])[0]
            
            start_dt, end_dt = parse_timeframe(timeframe, start_str, end_str)
            start_iso = start_dt.isoformat(timespec="seconds")
            end_iso = end_dt.isoformat(timespec="seconds")
            
            rows_dict = database.get_device_ping_logs(device_id, user["id"], start_iso, end_iso)
            total_checks = len(rows_dict)
            online_checks = sum(1 for r in rows_dict if r["status"] == "ONLINE")
            avg_health = round((online_checks / total_checks) * 100) if total_checks > 0 else 100
            
            aggregation = aggregate_pings(rows_dict, timeframe, start_dt, end_dt)
            
            today_hourly = None
            if timeframe == "today":
                hourly_labels = []
                for h in range(24):
                    if h == 0:
                        hourly_labels.append("12 AM")
                    elif h < 12:
                        hourly_labels.append(f"{h} AM")
                    elif h == 12:
                        hourly_labels.append("12 PM")
                    else:
                        hourly_labels.append(f"{h-12} PM")
                
                bucket_counts = {lbl: {"total": 0, "online": 0} for lbl in hourly_labels}
                for r in rows_dict:
                    try:
                        ts = datetime.fromisoformat(r["timestamp"])
                        h = ts.hour
                        lbl = hourly_labels[h]
                        bucket_counts[lbl]["total"] += 1
                        if r["status"] == "ONLINE":
                            bucket_counts[lbl]["online"] += 1
                    except Exception:
                        continue
                
                today_hourly = {
                    "labels": hourly_labels,
                    "uptime": [
                        round((bucket_counts[lbl]["online"] / bucket_counts[lbl]["total"]) * 100)
                        if bucket_counts[lbl]["total"] > 0 else None
                        for lbl in hourly_labels
                    ]
                }
            
            self.send_json({
                "avgHealth": avg_health,
                "overallAvg": aggregation["overallAvg"],
                "chartType": aggregation["chartType"],
                "labels": aggregation["labels"],
                "uptime": aggregation["uptime"],
                "latency": aggregation["latency"],
                "todayHourly": today_hourly
            })
            return

        # NEW EXCEL REPORT EXPORT ENDPOINT
        match = re.fullmatch(r"/api/companies/([^/]+)/analytics/export", path)
        if match:
            user = self.require_user()
            if not user:
                return
            company_id = unquote(match.group(1))
            if not find_company(user["id"], company_id):
                self.send_json({"error": "Company not found"}, 404)
                return
                
            timeframe = query.get("timeframe", ["week"])[0]
            start_str = query.get("start", [""])[0]
            end_str = query.get("end", [""])[0]
            device_id = query.get("device_id", [""])[0]
            
            start_dt, end_dt = parse_timeframe(timeframe, start_str, end_str)
            start_iso = start_dt.isoformat(timespec="seconds")
            end_iso = end_dt.isoformat(timespec="seconds")
            
            devices = company_devices(company_id)
            if device_id:
                devices = [d for d in devices if d["id"] == device_id]

            if not devices:
                self.send_json({"error": "No matching device found for this report."}, 404)
                return

            # ── Original simple Excel report format ────────────────────────────
            # Matches the plain layout: title line, period line, a bold
            # white-on-navy header row, plain data rows, and a bold light-green
            # average/summary row at the bottom. No charts, no side panels.
            wb = Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)

            HEADER_FILL  = "1F497D"
            AVERAGE_FILL = "E2EFDA"

            def _hdr_cell(cell, text):
                cell.value = text
                cell.font  = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                cell.fill  = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            def _avg_cell(cell, value):
                cell.value = value
                cell.font  = Font(name="Segoe UI", size=10, bold=True)
                cell.fill  = PatternFill(start_color=AVERAGE_FILL, end_color=AVERAGE_FILL, fill_type="solid")

            def fmt_seconds(total_seconds):
                total_seconds = int(round(total_seconds or 0))
                h, rem = divmod(total_seconds, 3600)
                m, s   = divmod(rem, 60)
                parts  = []
                if h: parts.append(f"{h}h")
                if m: parts.append(f"{m}m")
                if s or not parts: parts.append(f"{s}s")
                return " ".join(parts)

            def fmt_time(iso_str):
                try:
                    return datetime.fromisoformat(iso_str).strftime("%I:%M:%S %p")
                except Exception:
                    return iso_str or "-"

            # ── Per-device sheet ──────────────────────────────────────────────
            for dev in devices:
                sheet_title = clean_sheet_title(dev["name"], dev["location"])
                ws = wb.create_sheet(title=sheet_title)

                is_single_day = start_dt.date() == end_dt.date()

                if is_single_day:
                    data_headers = ["Session", "Start", "End", "Total Checks", "Unsuccessful", "Avg Latency (ms)", "Outage Time"]
                    col_widths   = [16, 13, 13, 13, 13, 17, 15]
                else:
                    data_headers = ["Date", "Total Checks", "Unsuccessful", "Avg Latency (ms)", "Total Outage Time", "Health %"]
                    col_widths   = [53, 15, 15, 19, 20, 12]
                n_data_cols = len(data_headers)

                for i, w in enumerate(col_widths, start=1):
                    ws.column_dimensions[get_column_letter(i)].width = w

                # Row 1 — title line
                ws["A1"].value = f"Health Report: {dev['name']} [{dev['location']}] ({dev['ip']})"
                ws["A1"].font  = Font(name="Calibri", size=11)

                # Row 2 — period line
                tf_label = timeframe.upper()
                ws["A2"].value = f"Period: {start_dt.strftime('%d-%m-%Y')} to {end_dt.strftime('%d-%m-%Y')} | Timeframe: {tf_label}"
                ws["A2"].font  = Font(name="Calibri", size=11)

                # Row 3 — blank spacer

                # Row 4 — data table header
                HDR_ROW = 4
                for ci, hdr in enumerate(data_headers, start=1):
                    _hdr_cell(ws.cell(row=HDR_ROW, column=ci), hdr)

                DATA_START_ROW = HDR_ROW + 1
                grand_total = grand_success = grand_failed = 0
                grand_outage_seconds = 0.0
                grand_latencies = []
                last_data_row   = DATA_START_ROW

                if is_single_day:
                    date_str   = start_dt.strftime("%Y-%m-%d")
                    day_stats  = database.get_day_stats(dev["id"], user["id"], date_str)
                    sessions   = day_stats["sessions"]

                    if not sessions:
                        ws.cell(row=DATA_START_ROW, column=1).value = "No monitoring sessions recorded for this day."
                        ws.cell(row=DATA_START_ROW, column=1).font = Font(name="Calibri", size=11)
                        last_data_row = DATA_START_ROW

                    for i, s in enumerate(sessions or [], start=1):
                        row_n = DATA_START_ROW + i - 1
                        vals = [
                            f"Session {i}",
                            fmt_time(s["start"]),
                            fmt_time(s["end"]),
                            s["total_checks"],
                            s["failed"],
                            f'{s["avg_latency_ms"]} ms' if s["avg_latency_ms"] is not None else "-",
                            fmt_seconds(s["outage_seconds"]),
                        ]
                        for ci, v in enumerate(vals, start=1):
                            cell = ws.cell(row=row_n, column=ci)
                            cell.value = v
                            cell.font  = Font(name="Calibri", size=11)
                        last_data_row = row_n

                    day_health = round((day_stats["successful"] / day_stats["total_checks"]) * 100) if day_stats["total_checks"] else None

                    avg_row = last_data_row + 2
                    _avg_cell(ws.cell(row=avg_row, column=1), "DAY AVERAGE")
                    _avg_cell(ws.cell(row=avg_row, column=4), day_stats["total_checks"])
                    _avg_cell(ws.cell(row=avg_row, column=5), day_stats["failed"])
                    _avg_cell(ws.cell(row=avg_row, column=6), f'{day_stats["avg_latency_ms"]} ms' if day_stats["avg_latency_ms"] is not None else "-")
                    _avg_cell(ws.cell(row=avg_row, column=7), fmt_seconds(day_stats["outage_seconds"]))

                else:
                    curr = datetime(start_dt.year, start_dt.month, start_dt.day)
                    row_n = DATA_START_ROW

                    while curr.date() <= end_dt.date():
                        date_str  = curr.strftime("%Y-%m-%d")
                        day_stats = database.get_day_stats(dev["id"], user["id"], date_str)
                        health    = round((day_stats["successful"] / day_stats["total_checks"]) * 100) if day_stats["total_checks"] else None

                        date_label = curr.strftime("%b %d (%a)")
                        vals = [
                            date_label,
                            day_stats["total_checks"],
                            day_stats["failed"],
                            f'{day_stats["avg_latency_ms"]} ms' if day_stats["avg_latency_ms"] is not None else "-",
                            fmt_seconds(day_stats["outage_seconds"]),
                            f"{health}%" if health is not None else "No data",
                        ]
                        for ci, v in enumerate(vals, start=1):
                            cell = ws.cell(row=row_n, column=ci)
                            cell.value = v
                            cell.font  = Font(name="Calibri", size=11)

                        grand_total            += day_stats["total_checks"]
                        grand_success          += day_stats["successful"]
                        grand_failed           += day_stats["failed"]
                        grand_outage_seconds   += day_stats["outage_seconds"] or 0
                        if day_stats["avg_latency_ms"] is not None:
                            grand_latencies.append(day_stats["avg_latency_ms"])
                        curr  += timedelta(days=1)
                        row_n += 1

                    overall_avg = round((grand_success / grand_total) * 100) if grand_total else None
                    overall_lat = round(sum(grand_latencies) / len(grand_latencies), 1) if grand_latencies else None
                    last_data_row = row_n - 1

                    avg_row = last_data_row + 2
                    avg_vals = [
                        f"{timeframe.upper()} AVERAGE", grand_total, grand_failed,
                        f"{overall_lat} ms" if overall_lat is not None else "-",
                        fmt_seconds(grand_outage_seconds),
                        f"{overall_avg}%" if overall_avg is not None else "No data",
                    ]
                    for ci, v in enumerate(avg_vals, start=1):
                        _avg_cell(ws.cell(row=avg_row, column=ci), v)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            excel_bytes = output.read()
            
            def sanitize_filename_part(text):
                return re.sub(r"[^a-zA-Z0-9_\-]", "_", str(text).strip())

            dev_name = sanitize_filename_part(devices[0]["name"]) if devices else "All_Devices"
            dev_loc = sanitize_filename_part(devices[0]["location"]) if devices else "Global"
            
            if timeframe == "today":
                date_str = start_dt.strftime("%Y-%m-%d")
                filename = f"Report_{dev_name}_{dev_loc}_{date_str}.xlsx"
            elif timeframe == "week":
                filename = f"Report_{dev_name}_{dev_loc}_Weekly_Report_{start_dt.strftime('%Y%m%d')}_to_{end_dt.strftime('%Y%m%d')}.xlsx"
            elif timeframe == "month":
                month_name = start_dt.strftime("%B_%Y")
                filename = f"Report_{month_name}_{dev_name}_{dev_loc}.xlsx"
            else: # custom
                filename = f"Report_{dev_name}_{dev_loc}_Custom_{start_dt.strftime('%Y%m%d')}_to_{end_dt.strftime('%Y%m%d')}.xlsx"

            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header(
                "Content-Disposition",
                f'attachment; filename="{filename}"'
            )
            self.send_header("Content-Length", str(len(excel_bytes)))
            self.end_headers()
            self.wfile.write(excel_bytes)
            return

        self.send_static(path)

    def do_POST(self):
        try:
            self._do_POST_inner()
        except (BrokenPipeError, ConnectionResetError):
            pass
        except Exception as exc:
            print(f"[error] POST {self.path} -> {exc}", file=sys.stderr)
            try:
                self.send_json({"error": "Something went wrong on the server. Please try again."}, 500)
            except Exception:
                pass

    def _do_POST_inner(self):
        path = urlparse(self.path).path

        if path == "/api/signup":
            data = self.read_json()
            name = str(data.get("name", "")).strip() or "User"
            email = str(data.get("email", "")).strip().lower()
            password = str(data.get("password", ""))
            company_name = str(data.get("company", "")).strip() or "My Company"
            username = str(data.get("username", "")).strip().lower()

            if not email or not password:
                self.send_json({"error": "Email and password are required"}, 400)
                return

            if get_user_by_email(email):
                self.send_json({"error": "Account already exists"}, 400)
                return

            if username and database.get_user_by_username(username):
                self.send_json({"error": "That user ID is already taken"}, 400)
                return

            try:
                user = database.create_user_record(name, email, password, company_name, username or None)
            except ValueError as exc:
                self.send_json({"error": str(exc)}, 400)
                return

            database.bootstrap_user_data(user["id"], company_name)
            companies = user_companies(user["id"])
            company = companies[0] if companies else None

            self.create_session(user)
            self.send_json({"user": public_user(user), "company": company}, 201)
            return

        if path == "/api/username-check":
            data = self.read_json()
            username = str(data.get("username", "")).strip().lower()
            self.send_json({"available": database.is_username_available(username)})
            return

        if path == "/api/username-suggest":
            data = self.read_json()
            name = str(data.get("name", ""))
            email = str(data.get("email", ""))
            self.send_json({"username": database.suggest_username(name, email)})
            return

        if path == "/api/login":
            data = self.read_json()
            identifier = str(data.get("identifier") or data.get("email") or "").strip().lower()
            password = str(data.get("password", ""))
            user = database.get_user_by_identifier(identifier)
            
            if not user or not verify_password(password, user["password_hash"]):
                self.send_json({"error": "Invalid user ID/email or password"}, 401)
                return
                
            self.create_session(user)
            self.send_json({"user": public_user(user)})
            return

        if path == "/api/logout":
            self.clear_session()
            self.send_json({"ok": True})
            return

        if path == "/api/forgot-password":
            data = self.read_json()
            email = str(data.get("email", "")).strip().lower()
            if not email:
                self.send_json({"error": "Enter your email address"}, 400)
                return

            user = get_user_by_email(email)
            if not user:
                self.send_json({"error": "This email is not registered"}, 404)
                return

            code = database.create_password_reset_code(user["id"])
            sent = False
            try:
                sent = email_alerts.send_password_reset_code(user["email"], code)
            except Exception as e:
                print(f"[Email] Failed to send password reset code: {e}")

            if not sent:
                self.send_json({"error": "Couldn't send the reset email right now. Please try again shortly."}, 502)
                return

            self.send_json({"ok": True, "message": "A reset code has been sent to your email."})
            return

        if path == "/api/reset-password":
            data = self.read_json()
            email = str(data.get("email", "")).strip().lower()
            code = str(data.get("code", "")).strip()
            new_password = str(data.get("new_password", ""))

            if not email or not code or not new_password:
                self.send_json({"error": "Email, code, and a new password are required"}, 400)
                return
            if len(new_password) < 6:
                self.send_json({"error": "New password must be at least 6 characters"}, 400)
                return

            user = get_user_by_email(email)
            if not user:
                self.send_json({"error": "This email is not registered"}, 404)
                return

            if not database.verify_and_consume_reset_code(user["id"], code):
                self.send_json({"error": "That code is invalid or has expired"}, 400)
                return

            database.reset_user_password(user["id"], new_password)
            self.send_json({"ok": True, "message": "Password updated. Please log in."})
            return

        user = self.require_user()
        if not user:
            return

        if path == "/api/companies":
            data = self.read_json()
            target_email = str(data.get("email", "")).strip()
            company = database.create_company(
                user["id"],
                str(data.get("name", "")).strip() or "New Company",
                target_email,
                target_email,
                30,
            )
                
            self.send_json({"company": company}, 201)
            return

        match = re.fullmatch(r"/api/companies/([^/]+)/devices", path)
        if match:
            company_id = unquote(match.group(1))
            if not find_company(user["id"], company_id):
                self.send_json({"error": "Company not found"}, 404)
                return
                
            device = clean_device(self.read_json(), company_id, user["id"])
            database.create_device(device)
                
            with STATE_LOCK:
                MONITOR_STATE.pop(device["id"], None)
                
            self.send_json({"device": device}, 201)
            return

        match = re.fullmatch(r"/api/companies/([^/]+)/devices/import", path)
        if match:
            company_id = unquote(match.group(1))
            if not find_company(user["id"], company_id):
                self.send_json({"error": "Company not found"}, 404)
                return

            data = self.read_json()
            try:
                devices, skipped = parse_csv_devices(str(data.get("csv", "")), company_id, user["id"])
            except Exception as exc:
                self.send_json({"error": f"Could not read that CSV file: {exc}"}, 400)
                return

            if not devices:
                self.send_json({"error": "No valid rows found — make sure the CSV has a Location, Name, and IP column, and at least one row with an IP address.", "skipped": skipped}, 400)
                return

            database.create_devices(devices)

            self.send_json({"imported": len(devices), "skipped": skipped})
            return

        # ── Incident Log: move one incident between Major and Minor ──
        match = re.fullmatch(r"/api/companies/([^/]+)/incidents/([^/]+)/reclassify", path)
        if match:
            company_id = unquote(match.group(1))
            incident_id = unquote(match.group(2))
            if not find_company(user["id"], company_id):
                self.send_json({"error": "Company not found"}, 404)
                return

            data = self.read_json()
            target = str(data.get("target", "")).lower()
            if target not in ("major", "minor"):
                self.send_json({"error": "target must be 'major' or 'minor'"}, 400)
                return

            if target == "major":
                if not incident_id.startswith("mint_"):
                    self.send_json({"error": "That incident is already Major."}, 400)
                    return
                minor = database.get_minor_interruption_by_id(incident_id)
                if not minor or minor.get("company_id") != company_id:
                    self.send_json({"error": "Incident not found"}, 404)
                    return
                result = database.reclassify_minor_to_major(incident_id)
                self.send_json({"success": True, "incident": result, "type": "major"})
                return
            else:
                if not incident_id.startswith("rep_"):
                    self.send_json({"error": "That incident is already Minor."}, 400)
                    return
                report = database.get_report_by_id(incident_id)
                if not report or report.get("company_id") != company_id:
                    self.send_json({"error": "Incident not found"}, 404)
                    return
                result = database.reclassify_report_to_minor(incident_id)
                self.send_json({"success": True, "incident": result, "type": "minor"})
                return

        self.send_json({"error": "Not found"}, 404)

    def do_PUT(self):
        try:
            self._do_PUT_inner()
        except (BrokenPipeError, ConnectionResetError):
            pass
        except Exception as exc:
            print(f"[error] PUT {self.path} -> {exc}", file=sys.stderr)
            try:
                self.send_json({"error": "Something went wrong on the server. Please try again."}, 500)
            except Exception:
                pass

    def _do_PUT_inner(self):
        user = self.require_user()
        if not user:
            return

        path = urlparse(self.path).path

        match = re.fullmatch(r"/api/companies/([^/]+)", path)
        if match:
            company_id = unquote(match.group(1))
            company = find_company(user["id"], company_id)
            if not company:
                self.send_json({"error": "Company not found"}, 404)
                return
                
            data = self.read_json()
            company["name"] = str(data.get("name", company["name"])).strip() or company["name"]
            
            update_email = str(data.get("receivers", company.get("receivers", ""))).strip()
            if update_email:
                company["email"] = update_email
                company["receivers"] = update_email
                
            company["alert_after_seconds"] = int(data.get("alert_after_seconds", company.get("alert_after_seconds", 30)))

            if "online_alerts" in data:
                company["online_alerts"] = 1 if data.get("online_alerts") else 0

            database.update_company(
                company_id,
                company["name"],
                company["email"],
                company["receivers"],
                company["alert_after_seconds"],
                company.get("online_alerts"),
            )
                
            self.send_json({"company": company})
            return

        match = re.fullmatch(r"/api/companies/([^/]+)/devices/([^/]+)", path)
        if match:
            company_id = unquote(match.group(1))
            device_id = unquote(match.group(2))
            
            if not find_company(user["id"], company_id):
                self.send_json({"error": "Company not found"}, 404)
                return
                
            device = database.get_device_in_company(device_id, company_id, user["id"])
            if not device:
                self.send_json({"error": "Device not found"}, 404)
                return
                
            updated = clean_device({**device, **self.read_json(), "id": device_id}, company_id, user["id"])
            database.update_device(
                device_id,
                company_id,
                user["id"],
                updated["location"],
                updated["name"],
                updated["ip"],
                updated["muted"],
            )
                
            with STATE_LOCK:
                MONITOR_STATE.pop(device_id, None)
                
            self.send_json({"device": updated})
            return

        self.send_json({"error": "Not found"}, 404)

    def do_DELETE(self):
        try:
            self._do_DELETE_inner()
        except (BrokenPipeError, ConnectionResetError):
            pass
        except Exception as exc:
            print(f"[error] DELETE {self.path} -> {exc}", file=sys.stderr)
            try:
                self.send_json({"error": "Something went wrong on the server. Please try again."}, 500)
            except Exception:
                pass

    def _do_DELETE_inner(self):
        user = self.require_user()
        if not user:
            return

        path = urlparse(self.path).path

        match = re.fullmatch(r"/api/companies/([^/]+)$", path)
        if match:
            company_id = unquote(match.group(1))
            if not find_company(user["id"], company_id):
                self.send_json({"error": "Company not found"}, 404)
                return
            if len(user_companies(user["id"])) <= 1:
                self.send_json({"error": "At least one company is required"}, 400)
                return
                
            database.delete_company(company_id)
                
            self.send_json({"ok": True})
            return

        match = re.fullmatch(r"/api/companies/([^/]+)/devices/([^/]+)", path)
        if match:
            company_id = unquote(match.group(1))
            device_id = unquote(match.group(2))
            
            if not find_company(user["id"], company_id):
                self.send_json({"error": "Company not found"}, 404)
                return
                
            if not database.get_device_in_company(device_id, company_id, user["id"]):
                self.send_json({"error": "Device not found"}, 404)
                return

            database.delete_device(device_id, company_id, user["id"])
            
            with STATE_LOCK:
                MONITOR_STATE.pop(device_id, None)
                
            self.send_json({"ok": True})
            return

        self.send_json({"error": "Not found"}, 404)

    def require_user(self):
        user = get_session_user(self)
        if not user:
            self.send_json({"error": "Unauthorized access"}, 401)
            return None
        return user

    def create_session(self, user):
        token = database.create_session(user["id"])
        self.extra_cookie = f"uptime_tools_session={token}; HttpOnly; Path=/; SameSite=Lax"

    def clear_session(self):
        cookie_header = self.headers.get("Cookie", "")
        cookie = SimpleCookie(cookie_header)
        token = cookie.get("uptime_tools_session")
        if token:
            database.delete_session(token.value)
        self.extra_cookie = "uptime_tools_session=; HttpOnly; Path=/; Max-Age=0; SameSite=Lax"

    def read_json(self):
        length = int(self.headers.get("Content-Length", "0"))
        raw_body = self.rfile.read(length).decode("utf-8")
        return json.loads(raw_body or "{}")

    def send_json(self, data, status=200):
        body = json.dumps(data).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        if hasattr(self, "extra_cookie"):
            self.send_header("Set-Cookie", self.extra_cookie)
            del self.extra_cookie
        self.end_headers()
        self.wfile.write(body)

    def send_static(self, request_path):
        if request_path == "/":
            request_path = "/index.html"
        
        clean_path = request_path.lstrip("/").replace("\\", "/")
        file_path = (WEB_DIR / clean_path).resolve()
        
        if not str(file_path).startswith(str(WEB_DIR.resolve())) or not file_path.exists():
            self.send_error(404)
            return
            
        content_type = mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
        body = file_path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format, *args):
        return


def device_monitor_loop(device_id, stop_event):
    """Runs for the lifetime of one device (until it's deleted or the
    process stops), completely independent of whether anyone has the
    dashboard open in a browser. Implements the full outage classification
    state machine:

      Normal monitoring -> ping every 5s.
      On failure -> verify with up to VERIFICATION_RETRY_COUNT 1s retries.
        Any retry succeeds  -> Temporary Packet Loss (no outage at all).
        All retries fail    -> confirmed outage, started at the timestamp
                                of the very first failed ping; switch to
                                1s monitoring until recovery.
      While still down -> the moment total downtime reaches
        MINOR_INTERRUPTION_THRESHOLD_SECONDS, an immediate "still offline"
        email is sent (once per outage) -- no need to wait for recovery.
      On recovery -> classify by duration:
        < MINOR_INTERRUPTION_THRESHOLD_SECONDS -> Minor Interruption
            (diagnostics + fluctuation detection only, never shown in the
            Incident Log / outage history, and no email).
        >= MINOR_INTERRUPTION_THRESHOLD_SECONDS -> Verified Outage (saved,
            reported, reduces availability, and a recovery email is sent).

    Exactly one thread ever pings a given device, so there's no race
    between this loop and the /status endpoint (which only reads cached
    state) or between two overlapping checks of the same device."""
    while not stop_event.is_set():
        device = database.get_device_by_id(device_id)
        if not device:
            return  # device was deleted -- stop monitoring it

        raw = ping_device(device)
        checked_at = datetime.now()

        if raw["raw_status"] == "ONLINE":
            record_online_check(device, raw, checked_at)
            if stop_event.wait(NORMAL_PING_INTERVAL_SECONDS):
                return
            continue

        # First failed ping -- begin verification (up to N 1s retries)
        first_failure_dt = checked_at
        record_verifying(device, 1, raw["ping"])

        recovered_during_verification = False
        last_raw = raw
        for attempt in range(1, VERIFICATION_RETRY_COUNT + 1):
            if stop_event.wait(VERIFICATION_RETRY_INTERVAL_SECONDS):
                return
            device = database.get_device_by_id(device_id)
            if not device:
                return
            last_raw = ping_device(device)
            if last_raw["raw_status"] == "ONLINE":
                recovered_during_verification = True
                break
            record_verifying(device, min(attempt + 1, VERIFICATION_RETRY_COUNT), last_raw["ping"])

        if recovered_during_verification:
            # Temporary Packet Loss: never became an outage at all.
            record_temporary_packet_loss(device)
            continue

        # All retries failed too -- confirmed outage.
        offline_start_iso = record_confirmed_outage_start(device, first_failure_dt)

        # Outage monitoring: ping every second until recovery.
        while not stop_event.is_set():
            if stop_event.wait(OUTAGE_PING_INTERVAL_SECONDS):
                return
            device = database.get_device_by_id(device_id)
            if not device:
                return
            outage_raw = ping_device(device)
            outage_checked_at = datetime.now()
            if outage_raw["raw_status"] == "ONLINE":
                record_recovery(device, offline_start_iso, outage_checked_at, outage_raw)
                break
            record_outage_still_down(device)

            # Still down -- fire the immediate "device is offline" email the
            # moment total downtime crosses the 60s threshold (once per outage).
            duration_so_far = (outage_checked_at - datetime.fromisoformat(offline_start_iso)).total_seconds()
            if duration_so_far >= MINOR_INTERRUPTION_THRESHOLD_SECONDS:
                record_long_outage_alert(device, offline_start_iso, outage_checked_at)
        # Recovered (or asked to stop) -- loop back to normal monitoring.


def device_supervisor_worker():
    """Runs in the background for the life of the process: keeps exactly
    one monitor thread alive per existing device, starting one the moment a
    device is added and stopping it the moment a device is deleted. Actual
    pinging happens inside device_monitor_loop, on each device's own
    schedule -- this just reconciles the roster periodically.

    Note: this only helps while the machine itself is powered on and awake.
    If the computer goes into OS-level sleep, every process on it --
    including this one -- is fully suspended, so no checks can run at all
    during that time. There's no software fix for that; the host machine
    needs to stay awake (or the app needs to run on something that's always
    on) for round-the-clock accuracy."""
    import time as _time

    _time.sleep(5)  # let startup settle first
    while True:
        try:
            live_device_ids = set()
            for company_id in database.all_company_ids():
                try:
                    for device in company_devices(company_id):
                        live_device_ids.add(device["id"])
                except Exception as e:
                    print(f"[monitor] Could not list devices for company {company_id}: {e}")

            with THREAD_REGISTRY_LOCK:
                # Start monitoring any device that doesn't have a thread yet.
                for device_id in live_device_ids:
                    entry = DEVICE_MONITOR_THREADS.get(device_id)
                    if entry and entry["thread"].is_alive():
                        continue
                    stop_event = threading.Event()
                    thread = threading.Thread(
                        target=device_monitor_loop, args=(device_id, stop_event), daemon=True
                    )
                    DEVICE_MONITOR_THREADS[device_id] = {"thread": thread, "stop_event": stop_event}
                    thread.start()

                # Stop monitoring (and forget) any device that no longer exists.
                for device_id in list(DEVICE_MONITOR_THREADS.keys()):
                    if device_id not in live_device_ids:
                        DEVICE_MONITOR_THREADS[device_id]["stop_event"].set()
                        del DEVICE_MONITOR_THREADS[device_id]
                        with STATE_LOCK:
                            MONITOR_STATE.pop(device_id, None)
        except Exception as e:
            print(f"[monitor] Roster reconciliation failed: {e}")

        _time.sleep(DEVICE_ROSTER_REFRESH_SECONDS)


def retention_worker():
    """Runs in the background for the life of the process: rolls up and
    prunes raw ping logs older than PING_LOG_RETENTION_DAYS. Runs shortly
    after startup, then every 6 hours. Also runs VACUUM roughly once a week
    (after the 28th rollup cycle) to reclaim disk space freed by pruning."""
    import time as _time

    _time.sleep(30)  # let startup settle first
    cycle = 0
    while True:
        try:
            database.rollup_and_prune_old_ping_logs(retention_days=PING_LOG_RETENTION_DAYS)
        except Exception as e:
            print(f"[retention] Rollup/prune failed: {e}")

        try:
            database.prune_old_minor_interruptions(retention_days=MINOR_INTERRUPTION_RETENTION_DAYS)
        except Exception as e:
            print(f"[retention] Minor interruption prune failed: {e}")

        cycle += 1
        if cycle % 28 == 0:  # 28 * 6h = ~1 week
            try:
                database.vacuum_database()
                print("[retention] VACUUM complete.")
            except Exception as e:
                print(f"[retention] VACUUM failed: {e}")

        _time.sleep(6 * 60 * 60)


def main():
    database.run_migrations()

    monitor_thread = threading.Thread(target=device_supervisor_worker, daemon=True)
    monitor_thread.start()

    retention_thread = threading.Thread(target=retention_worker, daemon=True)
    retention_thread.start()

    server = ThreadingHTTPServer((HOST, PORT), Handler)
    print(f"Uptime Tools is running at http://{HOST}:{PORT}")
    print("Press Ctrl+C to stop.")
    server.serve_forever()


if __name__ == "__main__":
    main()